from flask import Flask, render_template, session, redirect, url_for, request, flash, Response, send_file
from flask_sqlalchemy import SQLAlchemy
import os
from datetime import datetime
from werkzeug.utils import secure_filename
import requests
import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)

# ==========================================
# ⚙️ CONFIGURATION ULTRA-PRO (PRODUCTION)
# ==========================================
# Utilisation des variables d'environnement pour sécuriser les clés
app.secret_key = os.environ.get("SECRET_KEY", "flashpack_tunisia_2026_pro")
basedir = os.path.abspath(os.path.dirname(__file__))

UPLOAD_FOLDER = os.path.join(basedir, 'static', 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ---> CONNEXION POSTGRESQL <---
# Par défaut, utilise la base locale si aucune variable d'environnement n'est trouvée
DATABASE_URL = os.environ.get('DATABASE_URL', 'postgresql://postgres:ton_mot_de_passe@localhost:5432/flashpack_db')

# Correction automatique exigée par SQLAlchemy (postgres:// devient postgresql://)
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ==========================================
# 🗄️ MODÈLES DE DONNÉES (Base de données)
# ==========================================

class Produit(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text, nullable=False)
    details = db.Column(db.Text)
    prix_raw = db.Column(db.Float, nullable=False)
    prix_display = db.Column(db.String(20))
    image = db.Column(db.String(100))
    stock = db.Column(db.Integer, default=10)
    dimensions = db.Column(db.Text, nullable=True) # Ex: "10x15:0.500, 15x20:0.850"
    est_personnalisable = db.Column(db.Boolean, default=False)

class Client(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), nullable=False)
    tel = db.Column(db.String(20), nullable=False, unique=True) # Unique pour ne pas avoir de doublons
    adresse = db.Column(db.Text, nullable=False)
    mdp = db.Column(db.String(100), nullable=True) # Optionnel
    
    commandes = db.relationship('Commande', backref='client', lazy=True)

class Commande(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    date_commande = db.Column(db.DateTime, default=datetime.utcnow)
    details_panier = db.Column(db.Text)
    total_ttc = db.Column(db.Float)
    statut_livraison = db.Column(db.String(50), default="En attente")
    client_id = db.Column(db.Integer, db.ForeignKey('client.id'))

# --- INITIALISATION BDD ---
def initialiser_bdd():
    # On ajoute un try/except pour éviter que l'appli crashe au démarrage si la BDD n'est pas encore prête sur le serveur
    try:
        if Produit.query.first() is None:
            p1 = Produit(nom="Film Étirable", description="Film de protection haute performance, ultra-résistant pour la palettisation.", details="Épaisseur : 23 microns | Longueur : 300m", prix_raw=15.000, prix_display="15.000 DT", image="Film eritable.jpg")
            p2 = Produit(nom="Scotch d'Emballage", description="Ruban adhésif haute fixation. Ne se déchire pas.", details="Largeur : 48mm | Longueur : 100m", prix_raw=1.500, prix_display="1.500 DT", image="scotch.jpg")
            p3 = Produit(nom="Caisse Américaine", description="Carton double cannelure renforcé pour stockage longue durée.", details="Dimensions variables", prix_raw=2.500, prix_display="2.500 DT", image="caisse americaine double ondulation Kraft.jpeg", dimensions="60x40x40cm:2.500, 80x60x60cm:4.000")
            p4 = Produit(nom="DoyPack Personnalisé", description="Sachet souple avec fond stable pour produits alimentaires.", details="Fermeture Zip | Impression incluse", prix_raw=0.500, prix_display="0.500 DT", image="DoyPack.jpg", dimensions="10x15:0.500, 15x20:0.850, 16x27:1.200", est_personnalisable=True)
            p5 = Produit(nom="Sachet Zip", description="Petit sachet transparent avec fermeture par pression.", details="Matière : PEBD | 50 microns", prix_raw=0.100, prix_display="0.100 DT", image="SachetZip.png", dimensions="4x6:0.050, 10x15:0.100, 20x30:0.250")
            
            db.session.add_all([p1, p2, p3, p4, p5])
            db.session.commit()
            print("Données initiales injectées avec succès.")
    except Exception as e:
        print(f"Erreur lors de l'initialisation de la BDD : {e}")

# ==========================================
# 🛠️ FONCTIONS UTILITAIRES
# ==========================================

def get_produits_compatibles():
    tous_les_produits = Produit.query.all()
    dictionnaire_produits = {}
    for p in tous_les_produits:
        dictionnaire_produits[p.id] = {
            "id": p.id,
            "nom": p.nom,
            "description": p.description,
            "details": p.details,
            "prix_raw": p.prix_raw,
            "prix": p.prix_display, 
            "image": p.image,
            "dimensions": p.dimensions, 
            "est_personnalisable": p.est_personnalisable 
        }
    return dictionnaire_produits

# ==========================================
# 🌐 ROUTES DU SITE
# ==========================================

@app.route("/")
def home():
    return render_template("home.html", produits=get_produits_compatibles())

@app.route("/boutique")
def boutique():
    return render_template("boutique.html", produits=get_produits_compatibles())

@app.route("/produit/<int:produit_id>")
def detail_produit(produit_id):
    p = Produit.query.get_or_404(produit_id)
    
    liste_dimensions = []
    if p.dimensions:
        for item in p.dimensions.split(','):
            if ':' in item:
                d, pr = item.split(':')
                liste_dimensions.append({'taille': d.strip(), 'prix': float(pr.strip())})

    produit_compatible = {
        "id": p.id, 
        "nom": p.nom, 
        "description": p.description, 
        "details": p.details, 
        "prix_raw": p.prix_raw, 
        "prix": p.prix_display, 
        "image": p.image,
        "est_personnalisable": p.est_personnalisable,
        "stock": p.stock
    }
    return render_template("produit.html", produit=produit_compatible, produit_id=produit_id, liste_dimensions=liste_dimensions)

# --- ROUTES UTILISATEURS ---

@app.route("/compte", methods=['GET', 'POST'])
def compte():
    erreur = None
    if request.method == 'POST':
        tel = request.form.get('tel')
        mdp = request.form.get('mdp')

        # TEST ADMIN
        admin_pass = os.environ.get("ADMIN_PASSWORD", "flashpack2026")
        if tel == "admin" and mdp == admin_pass:
            session['admin_logged_in'] = True
            return redirect(url_for('admin'))

        # TEST CLIENT
        client = Client.query.filter_by(tel=tel).first()
        if client and client.mdp == mdp:
            session['nom'] = client.nom
            session['tel'] = client.tel
            session['adresse'] = client.adresse
            return redirect(url_for('boutique'))
        else:
            erreur = "Identifiants incorrects."

    return render_template("compte.html", erreur=erreur)

@app.route("/inscription_debut")
def inscription_debut():
    return render_template("inscription_form_complet.html")

@app.route("/finaliser_inscription", methods=['POST'])
def finaliser_inscription():
    nom = request.form.get('nom')
    tel = request.form.get('tel')
    adresse = request.form.get('adresse')
    mdp = request.form.get('mdp')

    client_existant = Client.query.filter_by(tel=tel).first()
    
    if client_existant:
        return render_template("inscription_form_complet.html", erreur="Ce numéro est déjà utilisé. Veuillez vous connecter.")

    nouveau_client = Client(nom=nom, tel=tel, adresse=adresse, mdp=mdp)
    db.session.add(nouveau_client)
    db.session.commit()

    session['nom'] = nom
    session['tel'] = tel
    session['adresse'] = adresse
    
    return redirect(url_for('boutique'))

@app.route("/logout")
def logout():
    session.clear() 
    return redirect(url_for('home'))

# --- ROUTES PANIER & COMMANDES ---

@app.route("/ajouter_au_panier/<int:id>", methods=['POST'])
def ajouter_au_panier(id):
    if not session.get('nom'):
        return redirect(url_for('compte'))
        
    produit = Produit.query.get_or_404(id)
    quantite = int(request.form.get('quantite', 1))
    
    dimension = request.form.get('dimension', '')
    design_option = request.form.get('design_option', '')
    
    fichier_nom = None
    if 'fichier_design' in request.files:
        fichier = request.files['fichier_design']
        if fichier.filename != '':
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            fichier_nom = f"{timestamp}_{secure_filename(fichier.filename)}"
            fichier.save(os.path.join(app.config['UPLOAD_FOLDER'], fichier_nom))
            
    if 'panier' not in session:
        session['panier'] = []
        
    panier = session['panier']
    
    article = {
        'id': produit.id,
        'nom': produit.nom,
        'image': produit.image,  
        'prix': produit.prix_raw,
        'quantite': quantite,
        'sous_total': produit.prix_raw * quantite,
        'dimension': dimension,
        'design_option': design_option,
        'fichier_design': fichier_nom
    }
    
    panier.append(article)
    session['panier'] = panier
    session.modified = True
    
    return redirect(url_for('panier'))

@app.route("/panier")
def panier():
    panier_session = session.get('panier', [])
    total_ht = sum(float(item.get('sous_total', 0)) for item in panier_session if isinstance(item, dict))
    
    tva = total_ht * 0.19
    total_ttc = total_ht + tva
    
    return render_template("panier.html", panier=panier_session, total_ht=total_ht, tva=tva, total_ttc=total_ttc)

@app.route("/retirer_du_panier/<int:index>")
def retirer_du_panier(index):
    if 'panier' in session:
        panier = session['panier']
        if 0 <= index < len(panier):
            panier.pop(index) 
            session['panier'] = panier
            session.modified = True
    return redirect(url_for('panier'))

@app.route("/vider_panier")
def vider_panier():
    session.pop('panier', None)
    return redirect(url_for('panier'))

@app.route("/valider_commande", methods=['POST'])
def valider_commande():
    
    def envoyer_notif_telegram(commande_id, client_nom, client_tel, details, total):
        # Tokens sécurisés par variables d'environnement
        TOKEN = os.environ.get("TELEGRAM_TOKEN", "8670391374:AAGo34UMU6zpJvf8eurtEN1xKVg3NfuVKYQ") 
        CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "7102761055")
        
        message = (
            f"🚨 NOUVELLE COMMANDE ! 🚨\n\n"
            f"📦 Commande #{commande_id}\n"
            f"👤 Client : {client_nom} (📞 {client_tel})\n"
            f"🛒 Articles : {details}\n"
            f"💰 Total : {total:.3f} DT\n\n"
            f"👉 Go sur le panel admin !"
        )
        url = f"https://api.telegram.org/bot{TOKEN}/sendMessage"
        try:
            requests.post(url, data={"chat_id": CHAT_ID, "text": message}, timeout=3)
        except Exception as e:
            print("Erreur d'envoi Telegram :", e)

    panier_session = session.get('panier', [])
    if not panier_session:
        return redirect(url_for('boutique'))
    
    nom_client = session.get('nom') or request.form.get('nom', 'Client Non Renseigné')
    tel_client = session.get('tel') or request.form.get('tel', '00000000')
    adresse_client = session.get('adresse') or request.form.get('adresse', 'Non Renseignée')

    client = Client.query.filter_by(tel=tel_client).first()
    if not client:
        client = Client(nom=nom_client, tel=tel_client, adresse=adresse_client)
        db.session.add(client)
        db.session.commit()
    else:
        client.adresse = adresse_client 
        db.session.commit()

    total_ht = 0
    liste_noms = []
    
    for item in panier_session:
        if isinstance(item, dict):
            total_ht += float(item.get('sous_total', 0))
            dim_str = f" ({item.get('dimension')})" if item.get('dimension') else ""
            des_str = f" [Design: {item.get('design_option')}]" if item.get('design_option') and item.get('design_option') != 'Aucun' else ""
            file_str = f" [Fichier: {item.get('fichier_design')}]" if item.get('fichier_design') else ""
            
            liste_noms.append(f"{item.get('quantite')}x {item.get('nom')}{dim_str}{des_str}{file_str}")
            
            produit_bdd = Produit.query.get(item.get('id'))
            if produit_bdd:
                produit_bdd.stock = produit_bdd.stock - int(item.get('quantite', 0))
            
    total_ttc = total_ht * 1.19
    details = " | ".join(liste_noms)
    
    nouvelle_commande = Commande(
        details_panier=details, 
        total_ttc=round(total_ttc, 3),
        client_id=client.id 
    )
    
    db.session.add(nouvelle_commande)
    db.session.commit()

    envoyer_notif_telegram(
        commande_id=nouvelle_commande.id,
        client_nom=nom_client,
        client_tel=tel_client,
        details=details,
        total=total_ttc
    )
    
    # Nettoyage de la session après l'achat
    for key in ['panier', 'remise', 'nom', 'tel', 'adresse']:
        session.pop(key, None)
    
    try:
        return render_template("confirmation.html")
    except:
        return f"<h1>✅ Commande #{nouvelle_commande.id} validée !</h1><p>Total : {round(total_ttc, 3)} DT</p><a href='/'>Retour à l'accueil</a>"

# ==========================================
# 🛡️ ESPACE ADMINISTRATION
# ==========================================

@app.route("/admin_logout")
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('home'))

@app.route("/admin")
def admin():
    if not session.get('admin_logged_in'):
        return redirect(url_for('compte'))
        
    toutes_les_commandes = Commande.query.order_by(Commande.date_commande.desc()).all()
    total_clients = Client.query.count()
    chiffre_affaires = sum(c.total_ttc for c in toutes_les_commandes)
    
    return render_template("admin.html", commandes=toutes_les_commandes, total_clients=total_clients, ca=chiffre_affaires, Produit=Produit)

@app.route("/admin/ajouter_produit", methods=['GET', 'POST'])
def admin_ajouter_produit():
    if not session.get('admin_logged_in'):
        return redirect(url_for('compte'))

    if request.method == 'POST':
        nouveau_p = Produit(
            nom=request.form.get('nom'),
            description=request.form.get('description'),
            details=request.form.get('details'),
            prix_raw=float(request.form.get('prix')),
            prix_display=f"{request.form.get('prix')} DT",
            image=request.form.get('image'),
            stock=int(request.form.get('stock')),
            dimensions=request.form.get('dimensions'),
            est_personnalisable=True if request.form.get('perso') else False
        )
        db.session.add(nouveau_p)
        db.session.commit()
        return redirect(url_for('admin'))

    return render_template("admin_produit_form.html", action="Ajouter", produit=None)

@app.route("/admin/modifier_produit/<int:id>", methods=['GET', 'POST'])
def admin_modifier_produit(id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('compte'))
    
    p = Produit.query.get_or_404(id)

    if request.method == 'POST':
        p.nom = request.form.get('nom')
        p.description = request.form.get('description')
        p.details = request.form.get('details')
        p.prix_raw = float(request.form.get('prix'))
        p.prix_display = f"{request.form.get('prix')} DT"
        p.image = request.form.get('image')
        p.stock = int(request.form.get('stock'))
        p.dimensions = request.form.get('dimensions')
        p.est_personnalisable = True if request.form.get('perso') else False
        db.session.commit()
        return redirect(url_for('admin'))

    return render_template("admin_produit_form.html", action="Modifier", produit=p)

@app.route("/admin/supprimer_produit/<int:id>", methods=['GET', 'POST'])
def admin_supprimer_produit(id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('compte'))
    
    produit_a_supprimer = Produit.query.get_or_404(id)
    db.session.delete(produit_a_supprimer)
    db.session.commit()
    return redirect(url_for('admin'))

@app.route("/admin/supprimer_commande/<int:id>", methods=['GET', 'POST'])
def admin_supprimer_commande(id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('compte'))
    
    commande_a_supprimer = Commande.query.get_or_404(id)
    db.session.delete(commande_a_supprimer)
    db.session.commit()
    return redirect(url_for('admin'))

@app.route("/update_statut/<int:id>/<nouveau_statut>")
def update_statut(id, nouveau_statut):
    if not session.get('admin_logged_in'):
        return redirect(url_for('compte'))
    
    commande = Commande.query.get(id)
    if commande:
        commande.statut_livraison = nouveau_statut
        db.session.commit()
    return redirect(url_for('admin'))

@app.route("/admin/export")
def export_commandes():
    if not session.get('admin_logged_in'):
        return redirect(url_for('compte'))
        
    commandes = Commande.query.all()
    produits = Produit.query.all()
    
    wb = openpyxl.Workbook()
    ws_cmd = wb.active
    ws_cmd.title = "Historique des Commandes"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill_cmd = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    
    headers_cmd = ['N° Commande', 'Date', 'Nom du Client', 'Téléphone', 'Adresse', 'Total TTC (DT)', 'Statut de Livraison', 'Détails des articles']
    ws_cmd.append(headers_cmd)
    
    for cell in ws_cmd[1]:
        cell.font = header_font
        cell.fill = header_fill_cmd
        cell.alignment = Alignment(horizontal="center")
        
    for c in commandes:
        date_cmd = c.date_commande.strftime('%d/%m/%Y %H:%M') if hasattr(c, 'date_commande') and c.date_commande else "Inconnue"
        ws_cmd.append([f"CMD-{c.id}", date_cmd, c.client.nom, c.client.tel, c.client.adresse, round(c.total_ttc, 3), c.statut_livraison, c.details_panier])
        
    for col in ws_cmd.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws_cmd.column_dimensions[column].width = min(max_length + 2, 60)

    ws_stock = wb.create_sheet("État des Stocks")
    ws_stock.append(['Produit', 'Stock Actuel', 'Prix Unitaire', 'Statut'])
    
    header_fill_stock = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    for cell in ws_stock[1]:
        cell.font = header_font
        cell.fill = header_fill_stock
        cell.alignment = Alignment(horizontal="center")
        
    for p in produits:
        statut = "⚠️ CRITIQUE" if p.stock < 5 else "✅ OK"
        ws_stock.append([p.nom, f"{p.stock} unités", p.prix_display, statut])
        
    for col in ws_stock.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws_stock.column_dimensions[column].width = max_length + 5

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    mois_actuel = datetime.now().strftime("%m_%Y")
    
    return send_file(excel_file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f"Rapport_FlashPack_{mois_actuel}.xlsx")

@app.route("/nettoyage_total")
def nettoyage_total():
    session.clear() 
    return "Le site est nettoyé ! <a href='/'>Retour à l'accueil</a>"

# ==========================================
# 🚀 LANCEMENT DU SERVEUR
# ==========================================
if __name__ == "__main__":
    with app.app_context():
        db.create_all()  
        initialiser_bdd() 
        
    # Bind universel pour la production (Hébergeurs)
    port = int(os.environ.get("PORT", 8000))
    app.run(host="0.0.0.0", port=port, debug=os.environ.get("FLASK_DEBUG", "True") == "True")